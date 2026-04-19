"""
BMW MSRP orchestration script.

Usage:
    python run.py [--output-dir /path/to/data]

Runs skill.js to scrape live BMW China prices, then writes/updates
data/BMW.xlsx with a new date column. Generates a price-change report
in report/ if any prices changed since the last run.
"""

import json
import os
import subprocess
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# ── Paths ──────────────────────────────────────────────────────────────────────
SKILL_DIR = Path(__file__).parent
PROJECT_DIR = SKILL_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
REPORT_DIR = PROJECT_DIR / "report"
BMW_FILE = DATA_DIR / "BMW.xlsx"
BRAND = "BMW"
TODAY = date.today().isoformat()  # e.g. "2026-04-19"

# ── Styles ─────────────────────────────────────────────────────────────────────
YELLOW = PatternFill("solid", fgColor="FFFF00")   # new trim
ORANGE = PatternFill("solid", fgColor="FFA500")   # price changed
STRIKE = Font(strikethrough=True, color="808080") # discontinued


def run_skill() -> list[dict]:
    """Run skill.js and return parsed JSON results."""
    result = subprocess.run(
        ["node", str(SKILL_DIR / "skill.js")],
        capture_output=True, text=True, encoding="utf-8"
    )
    if result.returncode != 0:
        sys.exit(f"skill.js failed:\n{result.stderr}")
    try:
        data = json.loads(result.stdout)
    except json.JSONDecodeError as e:
        sys.exit(f"Could not parse skill.js output: {e}\nOutput was:\n{result.stdout[:500]}")
    return data


def load_or_create_workbook():
    """Load existing BMW.xlsx or create a fresh one with headers."""
    if BMW_FILE.exists():
        return openpyxl.load_workbook(BMW_FILE)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = BRAND
    ws.append(["brand", "model", "trim"])
    return wb


def get_date_columns(ws) -> dict[str, int]:
    """Return {date_string: col_index} for all date columns (col 4+)."""
    dates = {}
    for col in range(4, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            dates[str(val)] = col
    return dates


def get_existing_rows(ws) -> dict[tuple, int]:
    """Return {(model, trim): row_index} for all data rows."""
    rows = {}
    for row in range(2, ws.max_row + 1):
        model = ws.cell(row, 2).value
        trim  = ws.cell(row, 3).value
        if model or trim:
            rows[(str(model or ""), str(trim or ""))] = row
    return rows


def write_results(scraped: list[dict]):
    DATA_DIR.mkdir(exist_ok=True)
    REPORT_DIR.mkdir(exist_ok=True)

    wb = load_or_create_workbook()
    ws = wb[BRAND]

    date_cols = get_date_columns(ws)
    existing_rows = get_existing_rows(ws)

    # Add today's date column if not present
    if TODAY not in date_cols:
        new_col = ws.max_column + 1
        ws.cell(1, new_col).value = TODAY
        date_cols[TODAY] = new_col
    today_col = date_cols[TODAY]

    # Find the most recent previous date column for change detection
    prev_col = None
    prev_date = None
    for d, c in sorted(date_cols.items()):
        if d < TODAY:
            prev_col = c
            prev_date = d

    price_changes = []
    scraped_keys = set()

    for item in scraped:
        model = item["model"]
        trim  = item["trim"]
        price = item["priceWan"]
        key   = (model, trim)
        scraped_keys.add(key)

        if key not in existing_rows:
            # New trim — append row and highlight
            row = ws.max_row + 1
            ws.cell(row, 1).value = BRAND
            ws.cell(row, 2).value = model
            ws.cell(row, 3).value = trim
            price_cell = ws.cell(row, today_col)
            price_cell.value = price
            price_cell.fill = YELLOW
            existing_rows[key] = row
        else:
            row = existing_rows[key]
            price_cell = ws.cell(row, today_col)
            price_cell.value = price

            # Detect price change vs previous date
            if prev_col:
                old_price = ws.cell(row, prev_col).value
                if old_price is not None and old_price != price:
                    price_cell.fill = ORANGE
                    price_changes.append({
                        "brand": BRAND,
                        "model": model,
                        "trim": trim,
                        "old_date": prev_date,
                        "old_price": old_price,
                        "new_date": TODAY,
                        "new_price": price,
                        "change": round(price - float(old_price), 2),
                    })

    # Mark discontinued trims (in existing rows but not in scraped data)
    for key, row in existing_rows.items():
        if key not in scraped_keys:
            for col in range(1, today_col):
                cell = ws.cell(row, col)
                cell.font = STRIKE
            ws.cell(row, today_col).value = None  # blank for this date

    # Auto-width columns
    for col in range(1, ws.max_column + 1):
        max_len = max(
            (len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    wb.save(BMW_FILE)
    print(f"Saved {BMW_FILE}  ({len(scraped)} trims, {len(scraped_keys)} unique)")

    # Write price change report
    if price_changes:
        report_path = REPORT_DIR / f"BMW_price_changes_{TODAY}.csv"
        import csv
        with open(report_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f, fieldnames=["brand","model","trim","old_date","old_price","new_date","new_price","change"]
            )
            writer.writeheader()
            writer.writerows(price_changes)
        print(f"Price change report: {report_path}  ({len(price_changes)} changes)")
    else:
        print("No price changes detected.")


def main():
    print("Running BMW MSRP skill...")
    scraped = run_skill()
    print(f"Scraped {len(scraped)} trim entries from bmw.com.cn")
    write_results(scraped)


if __name__ == "__main__":
    main()
