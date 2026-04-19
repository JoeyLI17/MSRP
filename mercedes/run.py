"""
Mercedes-Benz China MSRP orchestration script.

Usage:
    python run.py

Runs skill.js to fetch live Mercedes-Benz China prices, then writes/updates
data/Mercedes.xlsx with a new date column. Generates a price-change report
in report/ if any prices changed since the last run.
"""

import json
import os
import subprocess
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

SKILL_DIR  = Path(__file__).parent
PROJECT_DIR = SKILL_DIR.parent
DATA_DIR   = PROJECT_DIR / "data"
REPORT_DIR = PROJECT_DIR / "report"
MB_FILE    = DATA_DIR / "Mercedes.xlsx"
BRAND      = "Mercedes-Benz"
TODAY      = date.today().isoformat()

YELLOW = PatternFill("solid", fgColor="FFFF00")
ORANGE = PatternFill("solid", fgColor="FFA500")
STRIKE = Font(strikethrough=True, color="808080")


def run_skill() -> list[dict]:
    result = subprocess.run(
        ["node", str(SKILL_DIR / "skill.js")],
        capture_output=True, text=True, encoding="utf-8"
    )
    if result.returncode != 0:
        sys.exit(f"skill.js failed:\n{result.stderr}")
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError as e:
        sys.exit(f"Could not parse skill.js output: {e}\nOutput:\n{result.stdout[:500]}")


def load_or_create_workbook():
    if MB_FILE.exists():
        return openpyxl.load_workbook(MB_FILE)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = BRAND
    ws.append(["brand", "model", "trim"])
    return wb


def get_date_columns(ws) -> dict[str, int]:
    return {str(ws.cell(1, c).value): c
            for c in range(4, ws.max_column + 1)
            if ws.cell(1, c).value}


def get_existing_rows(ws) -> dict[tuple, int]:
    rows = {}
    for row in range(2, ws.max_row + 1):
        model = ws.cell(row, 2).value
        trim  = ws.cell(row, 3).value
        if model or trim:
            rows[(str(model or ""), str(trim or ""))] = row
    return rows


def write_results(scraped: list[dict]):
    DATA_DIR.mkdir(exist_ok=True)
    REPORT_DIR.mkdir(exist_ok=True)

    wb = load_or_create_workbook()
    ws = wb[BRAND]

    date_cols    = get_date_columns(ws)
    existing_rows = get_existing_rows(ws)

    if TODAY not in date_cols:
        new_col = ws.max_column + 1
        ws.cell(1, new_col).value = TODAY
        date_cols[TODAY] = new_col
    today_col = date_cols[TODAY]

    prev_col, prev_date = None, None
    for d, c in sorted(date_cols.items()):
        if d < TODAY:
            prev_col, prev_date = c, d

    price_changes = []
    scraped_keys  = set()

    for item in scraped:
        model = item["model"]
        trim  = item["trim"]
        price = item["priceWan"]
        key   = (model, trim)
        scraped_keys.add(key)

        if key not in existing_rows:
            row = ws.max_row + 1
            ws.cell(row, 1).value = BRAND
            ws.cell(row, 2).value = model
            ws.cell(row, 3).value = trim
            price_cell = ws.cell(row, today_col)
            price_cell.value = price
            price_cell.fill  = YELLOW
            existing_rows[key] = row
        else:
            row = existing_rows[key]
            price_cell = ws.cell(row, today_col)
            price_cell.value = price

            if prev_col:
                old_price = ws.cell(row, prev_col).value
                if old_price is not None and old_price != price:
                    price_cell.fill = ORANGE
                    price_changes.append({
                        "brand": BRAND, "model": model, "trim": trim,
                        "old_date": prev_date, "old_price": old_price,
                        "new_date": TODAY,     "new_price": price,
                        "change": round(price - float(old_price), 2),
                    })

    for key, row in existing_rows.items():
        if key not in scraped_keys:
            for col in range(1, today_col):
                ws.cell(row, col).font = STRIKE
            ws.cell(row, today_col).value = None

    for col in range(1, ws.max_column + 1):
        max_len = max(
            (len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1)),
            default=10
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)

    wb.save(MB_FILE)
    print(f"Saved {MB_FILE}  ({len(scraped)} trims, {len(scraped_keys)} unique)")

    if price_changes:
        import csv
        report_path = REPORT_DIR / f"Mercedes_price_changes_{TODAY}.csv"
        with open(report_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(
                f, fieldnames=["brand","model","trim","old_date","old_price","new_date","new_price","change"]
            )
            writer.writeheader()
            writer.writerows(price_changes)
        print(f"Price change report: {report_path}  ({len(price_changes)} changes)")
    else:
        print("No price changes detected.")


def main():
    print("Running Mercedes-Benz MSRP skill...")
    scraped = run_skill()
    print(f"Fetched {len(scraped)} models from mercedes-benz.com.cn")
    write_results(scraped)


if __name__ == "__main__":
    main()
